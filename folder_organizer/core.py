# core.py

import os
import zipfile
import shutil
import folder_organizer.helpers
from openpyxl import load_workbook

class Connect_Request_Details():

    _windows_illegal_name_trans_table = None

    def __init__(self, file_input):

        self.engagement = self.get_engagement_name(file_input)
        self.pathdict = self.generate_pathdict(file_input)

        return

    def get_engagement_name(self, filename, sheetname="Documents", name_position="A8"):

        wb = load_workbook(filename)

        sheet_ranges = wb[sheetname]

        engagement_name = sheet_ranges[name_position].value

        return engagement_name

    # REQUIRES: Request ID and Custom Groups are in the following columns in the xlsx input file
    # Request ID - B / Custom Group 1 - C / Custom Group 2 - D / Custom Group 3 - E
    def generate_pathdict(self, filename, sheetname="Documents", start_row=7):

        wb = load_workbook(filename)

        sheet_ranges = wb[sheetname]

        pathdict = dict()

        curr_row = start_row

        while True:

            curr_row_str = str(curr_row)
            ReqID = sheet_ranges['B'+curr_row_str].value

            if ReqID == None:
                break

            ReqID = int(ReqID)

            Groups = []

            Groups.append(sheet_ranges['C'+curr_row_str].value)
            Groups.append(sheet_ranges['D'+curr_row_str].value)
            Groups.append(sheet_ranges['E'+curr_row_str].value)

            midpath = ""

            for comp in Groups:
                if comp != None:
                    midpath = os.path.join(midpath, self._sanitize_group_name(comp))

            if ReqID not in pathdict:
                pathdict[ReqID] = midpath

            curr_row += 1

        pathdict[0] = ""

        return pathdict

    @classmethod
    def _sanitize_group_name(cls, groupname):
        """Replace bad characters and remove trailing dots from parts."""
        table = cls._windows_illegal_name_trans_table
        if not table:
            illegal = ':<>|"?*/\\'
            table = str.maketrans(illegal, '_' * len(illegal))
            cls._windows_illegal_name_trans_table = table
        groupname = groupname.translate(table)
        groupname = groupname.strip()
        return groupname


class ZipfileLongPaths(zipfile.ZipFile):

    def __init__(self, filename="NoName", date_time=(1980,1,1,0,0,0), requestdetailname="NoName"):

        # Build Engagement name and patch dict for organizing the files
        myreqclass = Connect_Request_Details(requestdetailname)

        self.engagement = myreqclass.engagement
        self.pathdict = myreqclass.pathdict

        zipfile.ZipFile.__init__(self, filename, date_time)

    # Strip individual folder/file names to remove leading/trailing whitespaces
    def _strip_arcname(self, arcname):

        (head, tail) = os.path.split(arcname)

        if head == "":
            return arcname

        else:
            return os.path.join(self._strip_arcname(head).strip(), tail.strip())

    # Remove redundant middle folders from the path
    def _remove_MiddleFolder(self, arcname, MiddleFolder):

        print("arcname_input: "+arcname)

        (head, tail) = os.path.split(arcname)

        print("head:" + head)
        print("tail:" + tail)

        if head == "":
            return arcname

        else:
            if tail == MiddleFolder:
                return head
            else:
                return os.path.join(self._remove_MiddleFolder(head, MiddleFolder), tail)

    # Prase the folder name to get the Connect Request Number
    # return 0 if the arcname does not start with a number
    def _get_requestno(self, arcname):

        def isint(string):
            try:
                int(string)
                return True
            except ValueError:
                return False

        requestno = arcname.split()[0]
        if isint(requestno):
            return int(requestno)
        else:
            return 0

    def _extract_member(self, member, targetpath, pwd):

        # targetpath: home dir to extract members
        targetpath = folder_organizer.helpers.winapi_path(targetpath).strip()

        """Extract the ZipInfo object 'member' to a physical
           file on the path targetpath.
        """
        if not isinstance(member, zipfile.ZipInfo):
            member = self.getinfo(member)

        # build the destination pathname, replacing
        # forward slashes to platform specific separators.
        arcname = member.filename.replace('/', os.path.sep)

        if os.path.altsep:
            arcname = arcname.replace(os.path.altsep, os.path.sep)
        # interpret absolute pathname as relative, remove drive letter or
        # UNC path, redundant separators, "." and ".." components.
        arcname = os.path.splitdrive(arcname)[1]
        invalid_path_parts = ('', os.path.curdir, os.path.pardir)
        arcname = os.path.sep.join(x for x in arcname.split(os.path.sep)
                                   if x not in invalid_path_parts)
        if os.path.sep == '\\':
            # filter illegal characters on Windows
            arcname = self._sanitize_windows_name(arcname, os.path.sep)


        # prevent dir names from having trailing whitespaces - JS
        arcname = self._strip_arcname(arcname)
        # remove redundant middle folders - JS
        arcname = self._remove_MiddleFolder(arcname, "Response Documents")

        reqno = self._get_requestno(arcname)

        targetpath = os.path.join(targetpath, self.pathdict[reqno])
        targetpath = os.path.join(targetpath, arcname)

        targetpath = os.path.normpath(targetpath)

        # Create all upper directories if necessary.
        upperdirs = os.path.dirname(targetpath)
        if upperdirs and not os.path.exists(upperdirs):
            os.makedirs(upperdirs)

        if member.is_dir():
            if not os.path.isdir(targetpath):
                os.mkdir(targetpath)
            return targetpath

        with self.open(member, pwd=pwd) as source, \
             open(targetpath, "wb") as target:
            shutil.copyfileobj(source, target)

        return targetpath

    def extract_ET_name(self):
        self.extractall(self.engagement+"/")


# zip_name = "Connect Document Export - 13.Apr.2020_Response Docs.zip"
# request_detail_name = 'Connect Request Details - 13.Apr.2020.xlsx'
#
# with ZipfileLongPaths(zip_name, 'r', request_detail_name) as zip:
#     zip.extract_ET_name()
